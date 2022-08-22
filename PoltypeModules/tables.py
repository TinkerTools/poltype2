import csv
import os
import sys
import productiondynamics as prod
import terminate as term
import numpy as np
import plots
import pandas as pd




def ComputeDynamicsTime(poltype,timeleft):
    """
    Intent:
    Input:
    Output:
    Referenced By: 
    Description: 
    """
    upper=timeleft/poltype.lowerperf
    lower=timeleft/poltype.upperperf
    return lower,upper


def KeyLists(poltype):
    """
    Intent:
    Input:
    Output:
    Referenced By: 
    Description: 
    """
    
    lambdakeylist=['Ele-Lambda','Vdw-Lambda','Restraint-Lambda']

    boxinfokeylist=['Total Atom Number','Average Box Size','Prod MD Ensemb','Prod MD Time','Prod MD Steps','Prod MD Arc File Space','Dynamic Writeout Frequency (ps)','Dynamic Time Step (fs)','Equil Time NPT','Equil Time NVT','Physio Counterions','Neut Counterions','Ligand Charge','Receptor Charge','Ligand Name','Receptor Name']
    energykeylist=[u'ΔGˢᵒˡᵛ',u'ΔGˢᵒˡᵛᵉʳʳ',u'ΔHˢᵒˡᵛ',u'ΔHˢᵒˡᵛᵉʳʳ',u'ΔSˢᵒˡᵛ',u'ΔSˢᵒˡᵛᵉʳʳ',u'ΔGᶜᵒᵐᵖᶜᵒʳʳ',u'ΔGᶜᵒᵐᵖᵘⁿᶜᵒʳʳ',u'ΔGᶜᵒᵐᵖᶜᵒʳʳᵉʳʳ',u'ΔHᶜᵒᵐᵖ',u'ΔHᶜᵒᵐᵖᵉʳʳ',u'ΔSᶜᵒᵐᵖ',u'ΔSᶜᵒᵐᵖᵉʳʳ',u'ΔGᵃⁿᵃᶜᵒᵐᵖᶜᵒʳʳ',u'ΔGᵇᶦⁿᵈᶜᵒʳʳ',u'ΔHᵇᶦⁿᵈ',u'ΔHᵇᶦⁿᵈᵉʳʳ',u'ΔSᵇᶦⁿᵈ',u'ΔSᵇᶦⁿᵈᵉʳʳ',u'ΔGˢᵒˡᵛᵉˡᵉ',u'ΔGˢᵒˡᵛᵛᵈʷ',u'ΔGᶜᵒᵐᵖᵉˡᵉ',u'ΔGᶜᵒᵐᵖᵛᵈʷ',u'ΔG',u'ΔGᵉˣᵖ',u'ΔGᵉˡᵉ',u'ΔGᵛᵈʷ',u'ΔGᵍᵃˢ',u'ΔGˢᵒˡ',u'ΔGᵉˡᵉᵍᵃˢ',u'ΔGᵉˡᵉˢᵒˡ',u'ΔGᵛᵈʷᵍᵃˢ',u'ΔGᵛᵈʷˢᵒˡ']
    freeenergykeylist=[u'ΔGˢᵒˡᵛ',u'ΔGˢᵒˡᵛᵉʳʳ',u'ΔGᶜᵒᵐᵖᶜᵒʳʳ',u'ΔGᶜᵒᵐᵖᵘⁿᶜᵒʳʳ',u'ΔGᶜᵒᵐᵖᶜᵒʳʳᵉʳʳ',u'ΔGᵃⁿᵃᶜᵒᵐᵖᶜᵒʳʳ',u'ΔGᵇᶦⁿᵈᶜᵒʳʳ',u'ΔGˢᵒˡᵛᵉˡᵉ',u'ΔGˢᵒˡᵛᵛᵈʷ',u'ΔGᶜᵒᵐᵖᵉˡᵉ',u'ΔGᶜᵒᵐᵖᵛᵈʷ',u'ΔG',u'ΔGᵉˣᵖ',u'ΔGᵉˡᵉ',u'ΔGᵛᵈʷ',u'ΔGᵍᵃˢ',u'ΔGˢᵒˡ',u'ΔGᵉˡᵉᵍᵃˢ',u'ΔGᵉˡᵉˢᵒˡ',u'ΔGᵛᵈʷᵍᵃˢ',u'ΔGᵛᵈʷˢᵒˡ']
    summarykeylist=[u'ΔGˢᵒˡᵛ',u'ΔGᶜᵒᵐᵖᶜᵒʳʳ',u'ΔGᵃⁿᵃᶜᵒᵐᵖᶜᵒʳʳ',u'ΔGᵇᶦⁿᵈᶜᵒʳʳ']
        
    if poltype.averageenergies==True:
        energykeylist.extend([u'ΔGᵇᶦⁿᵈᶜᵒʳʳᵃᵛᵍ',u'ΔGᵇᶦⁿᵈᶜᵒʳʳᵃᵛᵍᵉʳʳ',u'ΔHᵇᶦⁿᵈᵃᵛᵍ',u'ΔSᵇᶦⁿᵈᵃᵛᵍ'])
        freeenergykeylist.extend([u'ΔGᵇᶦⁿᵈᶜᵒʳʳᵃᵛᵍ',u'ΔGᵇᶦⁿᵈᶜᵒʳʳᵃᵛᵍᵉʳʳ'])
        summarykeylist.extend([u'ΔGᵇᶦⁿᵈᶜᵒʳʳᵃᵛᵍ'])
    keylist=[]
    keylist.extend(lambdakeylist)
    keylist.extend(boxinfokeylist)
    keylist.extend(energykeylist)
    keylist.extend(freeenergykeylist)
    keylist.extend(summarykeylist)
    return lambdakeylist,boxinfokeylist,energykeylist,freeenergykeylist,summarykeylist,keylist


def CSVWriter(poltype,tempname,progress=False):
    """
    Intent:
    Input:
    Output:
    Referenced By: 
    Description: 
    """
    with open(tempname, mode='w') as energy_file:
        energy_writer = csv.writer(energy_file, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
        grabbedfreeenergydict=poltype.masterdict['freeenergy']
        grabbedenergydict=poltype.masterdict['energy']
        grabbedboxinfodicts=poltype.masterdict['boxinfo']
        grabbeddictlist=[grabbedfreeenergydict,grabbedenergydict]
        for dic in grabbedboxinfodicts:
            grabbeddictlist.append(dic)   
        wrotecolumnheaderslist=[False,False,False,False]
        if poltype.binding==True:
            tablesummarylist=['Gibbs Free Energy Change Table','Enthalpy, Entropy, Gibbs Energy Change Table','Complexation Simulation Info Table','Solvation Simulation Info Table']
        else:
            tablesummarylist=['Gibbs Free Energy Change Table','Enthalpy, Entropy, Gibbs Energy Change Table','Solvation Simulation Info Table','Complexation Simulation Info Table']
        for dictidx in range(len(grabbeddictlist)):
            tabledict=grabbeddictlist[dictidx]
            allvaluerows=[]
            lentokeylist={} # if some molecules have more informatoin in tables (i.e. one uses gas phase another does not) then need to find max keylist and use as column header
            for path in tabledict.keys():
                table=tabledict[path]
                table={k: v for k, v in table.items() if v is not None}
                keylist=list(table.keys())
                lentokeylist[len(keylist)]=keylist
            maxkeylen=max(lentokeylist.keys())
            maxkeylist=lentokeylist[maxkeylen]

            for path in tabledict.keys():
                wrotecolumnheaders=wrotecolumnheaderslist[dictidx]
                table=tabledict[path]
                valuelist=[]
                for key in maxkeylist:
                    if key in table.keys():
                        valuelist.append(table[key])
                    else:
                        valuelist.append(None)
                for i in range(len(valuelist)):
                    value=valuelist[i]
                    if value!=None:
                        try:
                            float(value)
                            value=float(value)
                            value=round(value,3)
                            valuelist[i]=value
                        except:
                            pass
                        

                keyrowlist=['Name']
                keyrowlist.extend(maxkeylist)
                emptyline=[None]*len(keyrowlist)
                summaryline=[None]*len(keyrowlist)
                summary=tablesummarylist[dictidx]
                summaryline[0]=summary
                head, tail = os.path.split(path)
                valuerowlist=[tail]
                valuerowlist.extend(valuelist)
                if wrotecolumnheaders==False:
                    energy_writer.writerow(emptyline)
                    energy_writer.writerow(summaryline)
                    energy_writer.writerow(keyrowlist)
                    wrotecolumnheaderslist[dictidx]=True
                energy_writer.writerow(valuerowlist)
                allvaluerows.append(valuerowlist) 
            WriteOutTable(poltype,allvaluerows,summary,keyrowlist)
        
    GenerateExcelFile(poltype)


def WriteOutTable(poltype,allvaluerows,summary,keyrowlist):
    """
    Intent:
    Input:
    Output:
    Referenced By: 
    Description: 
    """
    tempname=summary.replace(' ','_')+'.csv'
    with open(tempname, mode='w') as energy_file:
        energy_writer = csv.writer(energy_file, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
        energy_writer.writerow(keyrowlist)
        for row in allvaluerows:
            energy_writer.writerow(row) 



def GenerateSimInfoTable(poltype):
    """
    Intent:
    Input:
    Output:
    Referenced By: 
    Description: 
    """
    
    curdir=os.getcwd()
    os.chdir(poltype.outputpath)
    lambdakeylist,boxinfokeylist,energykeylist,freeenergykeylist,summarykeylist,keylist=KeyLists(poltype)
    OrderTableData(poltype,boxinfokeylist,lambdakeylist,energykeylist,freeenergykeylist,summarykeylist,poltype.outputpath)  
    if poltype.binding==True and poltype.complexationonly==False:
        EnterBindData(poltype,poltype.outputpath)      
    plots.PlotBARConvergence(poltype) 
    tempname='SimData.csv'
    CSVWriter(poltype,tempname,True)
    os.chdir(curdir)


def OrderTableData(poltype,boxinfokeylist,lambdakeylist,energykeylist,freeenergykeylist,summarykeylist,path):
    """
    Intent:
    Input:
    Output:
    Referenced By: 
    Description: 
    """
    for i in range(len(poltype.masterdict['boxinfo'])):
        if path not in poltype.masterdict['boxinfo'][i].keys():
            poltype.masterdict['boxinfo'][i][path]={}
    for i in range(len(poltype.tabledict)):
        table=poltype.tabledict[i]

        for key in boxinfokeylist:
            if key in table.keys():
                poltype.masterdict['boxinfo'][i][path][key]=table[key]
            else:
                if key not in poltype.masterdict['boxinfo'][i][path].keys():
                    poltype.masterdict['boxinfo'][i][path][key]=None

    if path not in poltype.masterdict['lambda'].keys():
        poltype.masterdict['lambda'][path]={}
    for i in range(len(poltype.tabledict)):
        table=poltype.tabledict[i]
        for key in lambdakeylist:
            if key in table.keys():
                poltype.masterdict['lambda'][path][key]=table[key]
            else:
                if key not in poltype.masterdict['lambda'][path].keys():
                    poltype.masterdict['lambda'][path][key]=None

    if path not in poltype.masterdict['energy'].keys():
        poltype.masterdict['energy'][path]={}

    for i in range(len(poltype.tabledict)):
        table=poltype.tabledict[i]
        for key in energykeylist:
            if key in table.keys():
                poltype.masterdict['energy'][path][key]=table[key]
            else:
                if key not in poltype.masterdict['energy'][path].keys():
                    poltype.masterdict['energy'][path][key]=None

    if path not in poltype.masterdict['freeenergy'].keys():
        poltype.masterdict['freeenergy'][path]={}
    for i in range(len(poltype.tabledict)):
        table=poltype.tabledict[i]
        for key in freeenergykeylist:
            if key in table.keys():
                poltype.masterdict['freeenergy'][path][key]=table[key]
            else:
                if key not in poltype.masterdict['freeenergy'][path].keys():
                    poltype.masterdict['freeenergy'][path][key]=None

    if path not in poltype.masterdict['summary'].keys():
        poltype.masterdict['summary'][path]={}
    for i in range(len(poltype.tabledict)):
        table=poltype.tabledict[i]
        for key in summarykeylist:
            if key in table.keys():
                poltype.masterdict['summary'][path][key]=table[key]
            else:
                if key not in poltype.masterdict['summary'][path].keys():
                    poltype.masterdict['summary'][path][key]=None


def GenerateFreeEnergyMatrix(poltype,ligandnames,receptornames,bindenergies):
    """
    Intent:
    Input:
    Output:
    Referenced By: 
    Description: 
    """
    uniquereceptors=len(set(receptornames))
    uniqueligands=len(ligandnames)
    mat=np.array((uniquereceptors,uniqueligands),dtype=float)
    receptortoligands={}
    receptorligandtoenergy={}
    for i in range(len(receptornames)):
        receptorname=receptornames[i]
        ligandname=ligandnames[i] 
        free=bindenergies[i]
        if receptorname not in receptortoligands.keys():
            receptortoligands[receptorname]=[]
        receptortoligands[receptorname].append(ligandname)
        if receptorname not in receptorligandtoenergy.keys():
            receptorligandtoenergy[receptorname]={}
        receptorligandtoenergy[receptorname][ligandname]=free
    count=0
    yaxis=[]
    xaxis=[]
    for receptor,ligands in receptortoligands.items(): 
        yaxis.append(receptor)
        for i in range(len(ligands)):
            ligand=ligands[i]
            energy=receptorligandtoenergy[receptor][ligand]
            if len(ligands)>1:
                mat[count,i]=float(energy)
            else:
                mat[count]=float(energy)

            xaxis.append(ligand)
            
        count+=1
    mat=np.transpose(mat)
    return mat,xaxis,yaxis

def GrabSimDataFromPathList(poltype):
    """
    Intent:
    Input:
    Output:
    Referenced By: 
    Description: 
    """
    
    topdir=os.getcwd()
    poltype.simpathlist=[]
    for root, subdirs, files in os.walk(poltype.pathtosims):
        for d in subdirs:
            curdir=os.getcwd()
            path=os.path.join(root, d)
            os.chdir(path)
            files=os.listdir()
            for f in files:
                if 'SimData.csv' in f:
                    poltype.simpathlist.append(path)
    lambdakeylist,boxinfokeylist,energykeylist,freeenergykeylist,summarykeylist,keylist=KeyLists(poltype)
    ligandnames=[]
    receptornames=[]
    bindenergies=[]
    for path in poltype.simpathlist:
        os.chdir(path)
        with open('SimData.csv') as csv_file:
            rows=[]
            csv_reader = csv.reader(csv_file, delimiter=',')
            for row in csv_reader:
                rows.append(row)
            for k in range(len(poltype.tabledict)):
                tabledict={}
                for i in range(len(rows)):
                    row=rows[i]
                    for j in range(len(row)):
                        col=row[j]
                        if col in keylist:
                            nextrow=rows[i+1]
                            colvalue=nextrow[j]
                            if colvalue!=None:
                                try:
                                    float(colvalue)
                                    colvalue=float(colvalue)
                                    colvalue=round(colvalue,3)
                                except:
                                    pass


                            tabledict[col]=colvalue
                poltype.tabledict[k]=tabledict
               
        OrderTableData(poltype,boxinfokeylist,lambdakeylist,energykeylist,freeenergykeylist,summarykeylist,path)  
        if poltype.binding==True:
            EnterBindData(poltype,path) 
        ligname=poltype.masterdict['boxinfo'][0][path]['Ligand Name']
        receptorname=poltype.masterdict['boxinfo'][0][path]['Receptor Name']
        freeenergy=poltype.masterdict['energy'][path][u'ΔGᵇᶦⁿᵈᶜᵒʳʳ']
        ligandnames.append(ligname)
        receptornames.append(receptorname)
        bindenergies.append(freeenergy) 

    os.chdir(topdir)
    if poltype.binding==True:
        mat,xaxis,yaxis=GenerateFreeEnergyMatrix(poltype,ligandnames,receptornames,bindenergies) 
        plots.PlotHeatmap(poltype,mat,xaxis,yaxis)


    if poltype.averageenergies==True:
        delGbindmodelist=[]
        delGbindmodelisterr=[]
        delHbindmodelist=[]
        delSbindmodelist=[]
        for path in poltype.masterdict['energy'].keys():
            if poltype.masterdict['energy'][path][u'ΔGᵇᶦⁿᵈᶜᵒʳʳ']!=None:
                delGbindmodelist.append(poltype.masterdict['energy'][path][u'ΔGᵇᶦⁿᵈᶜᵒʳʳ'])
            if poltype.masterdict['energy'][path][u'ΔGᵇᶦⁿᵈᶜᵒʳʳᵉʳʳ']!=None:
                delGbindmodelisterr.append(poltype.masterdict['energy'][path][u'ΔGᵇᶦⁿᵈᶜᵒʳʳᵉʳʳ'])
            if poltype.masterdict['energy'][path][u'ΔHᵇᶦⁿᵈ']!=None:
                delHbindmodelist.append(poltype.masterdict['energy'][path][u'ΔHᵇᶦⁿᵈ'])
            if poltype.masterdict['energy'][path][u'ΔSᵇᶦⁿᵈ']!=None:
                delSbindmodelist.append(poltype.masterdict['energy'][path][u'ΔSᵇᶦⁿᵈ'])

        AverageEnergyList(poltype,delGbindmodelist,'G')
        AverageEnergyList(poltype,delHbindmodelist,'H')
        AverageEnergyList(poltype,delSbindmodelist,'S')
        poltype.deltaGaverageerr=TotalBindingAffinityError(poltype,delGbindmodelist,delGbindmodelisterr)
        for path in poltype.masterdict['energy'].keys():
            poltype.masterdict['freeenergy'][path][u'ΔGᵇᶦⁿᵈᶜᵒʳʳᵃᵛᵍ']=poltype.deltaGaverage
            poltype.masterdict['energy'][path][u'ΔGᵇᶦⁿᵈᶜᵒʳʳᵃᵛᵍ']=poltype.deltaGaverage
            poltype.masterdict['freeenergy'][path][u'ΔGᵇᶦⁿᵈᶜᵒʳʳᵃᵛᵍᵉʳʳ']=poltype.deltaGaverageerr
            poltype.masterdict['energy'][path][u'ΔGᵇᶦⁿᵈᶜᵒʳʳᵃᵛᵍᵉʳʳ']=poltype.deltaGaverageerr
            poltype.masterdict['summary'][path][u'ΔGᵇᶦⁿᵈᶜᵒʳʳᵃᵛᵍ']=poltype.deltaGaverage
            poltype.masterdict['energy'][path][u'ΔHᵇᶦⁿᵈᵃᵛᵍ']=poltype.deltaHaverage
            poltype.masterdict['energy'][path][u'ΔSᵇᶦⁿᵈᵃᵛᵍ']=poltype.deltaSaverage


    tempname='GrabbedSimData.csv'
    CSVWriter(poltype,tempname)   
    plots.PlotFreeEnergyVsExp(poltype)
    return

def EnterBindData(poltype,path):
    """
    Intent:
    Input:
    Output:
    Referenced By: 
    Description: 
    """
    poltype.masterdict['summary'][path][u'ΔGˢᵒˡᵛ']=poltype.masterdict['freeenergy'][path][u'ΔGˢᵒˡᵛ']
    poltype.masterdict['summary'][path][u'ΔGᶜᵒᵐᵖᶜᵒʳʳ']=poltype.masterdict['freeenergy'][path][u'ΔGᶜᵒᵐᵖᶜᵒʳʳ']
    DelGBind=float(poltype.masterdict['energy'][path][u'ΔGᶜᵒᵐᵖᶜᵒʳʳ'])-float(poltype.masterdict['energy'][path][u'ΔGˢᵒˡᵛ'])
    DelSBind=float(poltype.masterdict['energy'][path][u'ΔSᶜᵒᵐᵖ'])-float(poltype.masterdict['energy'][path][u'ΔSˢᵒˡᵛ'])
    DelHBind=float(poltype.masterdict['energy'][path][u'ΔHᶜᵒᵐᵖ'])-float(poltype.masterdict['energy'][path][u'ΔHˢᵒˡᵛ'])
    solvGerr=float(poltype.masterdict['energy'][path][u'ΔGˢᵒˡᵛᵉʳʳ'])
    compGerr=float(poltype.masterdict['energy'][path][u'ΔGᶜᵒᵐᵖᶜᵒʳʳᵉʳʳ'])
    DelGBinderr=np.sqrt(solvGerr**2+compGerr**2)
    solvHerr=float(poltype.masterdict['energy'][path][u'ΔHˢᵒˡᵛᵉʳʳ'])
    compHerr=float(poltype.masterdict['energy'][path][u'ΔHᶜᵒᵐᵖᵉʳʳ'])
    DelHBinderr=np.sqrt(solvHerr**2+compHerr**2)
    solvSerr=float(poltype.masterdict['energy'][path][u'ΔSˢᵒˡᵛᵉʳʳ'])
    compSerr=float(poltype.masterdict['energy'][path][u'ΔSᶜᵒᵐᵖᵉʳʳ'])
    DelSBinderr=np.sqrt(solvSerr**2+compSerr**2)
    poltype.masterdict['energy'][path][u'ΔGᵇᶦⁿᵈᶜᵒʳʳ']=str(DelGBind)
    poltype.masterdict['energy'][path][u'ΔGᵇᶦⁿᵈᶜᵒʳʳᵉʳʳ']=str(DelGBinderr)
    poltype.masterdict['energy'][path][u'ΔHᵇᶦⁿᵈ']=str(DelHBind)
    poltype.masterdict['energy'][path][u'ΔHᵇᶦⁿᵈᵉʳʳ']=str(DelHBinderr)
    poltype.masterdict['energy'][path][u'ΔSᵇᶦⁿᵈ']=str(DelSBind)
    poltype.masterdict['energy'][path][u'ΔSᵇᶦⁿᵈᵉʳʳ']=str(DelSBinderr)
    poltype.masterdict['freeenergy'][path][u'ΔGᵇᶦⁿᵈᶜᵒʳʳ']=str(DelGBind)
    poltype.masterdict['summary'][path][u'ΔGᵇᶦⁿᵈᶜᵒʳʳ']=str(DelGBind)
    poltype.masterdict['freeenergy'][path][u'ΔGᵇᶦⁿᵈᶜᵒʳʳᵉʳʳ']=str(DelGBinderr)
    poltype.masterdict['energy'][path][u'ΔGᵇᶦⁿᵈᶜᵒʳʳ']=str(DelGBind)
    poltype.masterdict['energy'][path][u'ΔGᵇᶦⁿᵈᶜᵒʳʳᵉʳʳ']=str(DelGBinderr)
    poltype.masterdict['energy'][path][u'ΔHᵇᶦⁿᵈ']=str(DelHBind)
    poltype.masterdict['energy'][path][u'ΔHᵇᶦⁿᵈᵉʳʳ']=str(DelHBinderr)
    poltype.masterdict['energy'][path][u'ΔSᵇᶦⁿᵈ']=str(DelSBind)
    poltype.masterdict['energy'][path][u'ΔSᵇᶦⁿᵈᵉʳʳ']=str(DelSBinderr)
    poltype.masterdict['freeenergy'][path][u'ΔGᵇᶦⁿᵈᶜᵒʳʳ']=str(DelGBind)
    poltype.masterdict['summary'][path][u'ΔGᵇᶦⁿᵈᶜᵒʳʳ']=str(DelGBind)
    poltype.masterdict['freeenergy'][path][u'ΔGᵇᶦⁿᵈᶜᵒʳʳᵉʳʳ']=str(DelGBinderr)


def AverageEnergyList(poltype,enlist,key):
    """
    Intent:
    Input:
    Output:
    Referenced By: 
    Description: 
    """
    if 'G' in key:
        poltype.deltaGaverage=TotalBindingAffinity(poltype,enlist)
    elif 'H' in key:
        poltype.deltaHaverage=BoltzmannAverage(poltype,enlist)
    elif 'S' in key:
        poltype.deltaSaverage=(poltype.deltaHaverage-poltype.deltaGaverage)/int(poltype.equilibriatescheme[-1])

def TotalBindingAffinity(poltype,enlist):
    """
    Intent:
    Input:
    Output:
    Referenced By: 
    Description: 
    """
    kB=0.0019872041 # kcal/molK
    T=int(poltype.equilibriatescheme[-1])
    sumboltzfactors=0
    for obs in enlist:
        sumboltzfactors+=np.exp(-obs/(kB*T))
    totalsum=-kB*T*np.log(sumboltzfactors)
    return totalsum

def TotalBindingAffinityError(poltype,enlist,errlist):
    """
    Intent:
    Input:
    Output:
    Referenced By: 
    Description: 
    """
    kB=0.0019872041 # kcal/molK
    T=int(poltype.equilibriatescheme[-1])
    Sum=0
    for obsidx in range(len(enlist)):
        obs=enlist[obsidx]
        err=errlist[obsidx]
        Sum+=(err/obs)**2
    return np.sqrt(Sum)

                
def BoltzmannAverage(poltype,enlist):
    """
    Intent:
    Input:
    Output:
    Referenced By: 
    Description: 
    """
    kB=0.0019872041
    T=int(poltype.equilibriatescheme[-1])
    sumboltzfactors=0
    for obs in enlist:      
        sumboltzfactors+=np.exp(-obs/(kB*T))
    boltzavg=0
    for obs in enlist:
        boltzfactor=np.exp(-obs/(kB*T))
        boltzavg+=obs*boltzfactor
    boltzavg=boltzavg/sumboltzfactors
    return boltzavg  


def WriteTableUpdateToLog(poltype,verbose=True):
    """
    Intent:
    Input:
    Output:
    Referenced By: 
    Description: 
    """
    for i in range(len(poltype.tabledict)):
        table=poltype.tabledict[i]
        tableused=poltype.tabledictkeysused[i]
        for key in table.keys():
            value=table[key]
            if key not in tableused:
                poltype.tabledictkeysused[i].append(key)
                if verbose==True:
                    poltype.WriteToLog(key+' = '+str(value))


def GenerateExcelFile(poltype):
    from openpyxl.styles import Color, PatternFill, Font, Border
    from openpyxl.styles import colors
    from openpyxl.cell import Cell
    from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
    from openpyxl.utils import get_column_letter
    spreadsheet='SimulationResults.xlsx'
    files=os.listdir()
    csvfiles=[]
    for f in files:
        if ".csv" in f and 'SimData' not in f:
            csvfiles.append(f)

    with pd.ExcelWriter(spreadsheet, engine="openpyxl") as writer:
        for f in csvfiles:
            df = pd.read_csv(f)
            split=f.split('.csv')
            sheet_name = split[0]
            df.to_excel(writer, sheet_name=sheet_name)
            sheet = writer.sheets[sheet_name]
            column_widths = []
            for row in sheet.iter_rows():
                for i, cell in enumerate(row):
                    try:
                        column_widths[i] = max(column_widths[i], len(str(cell.value)))
                    except IndexError:
                        column_widths.append(len(str(cell.value)))
            
            for i, column_width in enumerate(column_widths):
                sheet.column_dimensions[get_column_letter(i + 1)].width = column_width

